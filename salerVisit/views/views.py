from django.shortcuts import render
from io import BytesIO
from mandala.auth.decorators import permission_required,login_required
from django.http import JsonResponse, HttpResponse
from public.utils import get_value_list
from public.Time import Time
from public.config import problem_content,problem_content1
from customer.models import SalesPortraitActivity,User as CrmUser
from salerVisit.forms import VisitSelectForm
from salerVisit.utils import get_groups,get_department
from secretary.models import DingGroupMemberMap
from datetime import datetime,timedelta
import pandas as pd
import numpy as np
import json
from openpyxl import Workbook
# Create your views here.
from user_center.models import UserLog
from django.urls import reverse
# from support import settings
from notice.models import NoticeList
from secretary.utils import get_departments_about_sale, get_group_members
from django.conf import settings
URL_403 = settings.URL_403
JSON_403 = settings.JSON_403

'''全国签到页面'''
@login_required()
@permission_required("salerVisit.visit_detail.view",login_url=URL_403)
def visitHtml(request):
    visitDetailData = visitCountMethod(request)
    visit_count = visitDetailData.get('visit_count')
    group_count = visitDetailData.get('group_count')
    end_time = visitDetailData.get('end_time')
    visit_detail = visitDetailData.get('visit_detail')
    ########## 跳转CRM客户名称首页 #############
    crm_customer_jump_url = settings.CRM_CUSTOMER_JUMP_URL
    # print(visit_count)
    # print(visit_detail)
    # print(group_count)
    # print(end_time)
    title = "全国签到"
    user = request.user
    if user.username not in ["武翠霞", "丛大侠", "王寅", "张太锋", "黄远", "朱铭鑫", "袁战梅"]:
        message = json.dumps(dict(request.GET))
        UserLog.objects.create(username=user.username, user_id=user.id,
                               model="全国签到", action="进入页面", message=message)
    return render(request, 'salerVisit/map.html', locals())


'''计算今日/昨日/本周/上月/本月/本年度签到次数,
    参数为城市id，如果没有则是全国'''
def visitCountMethod(request):
    result = {"status": 0, "message": ""}
    # params = request.GET.get()
    time_obj = Time()
    end_time = time_obj.now.strftime('%Y-%m-%d %H:00')
    # last_hours_start = time_obj.last_hours().strftime('%Y-%m-%d %H:')
    today = time_obj.totay().strftime('%Y-%m-%d')
    yestoday = time_obj.yesterday().strftime('%Y-%m-%d')
    this_week_start = time_obj.this_week_start().strftime('%Y-%m-%d 00:00')
    this_month_start = time_obj.this_month_start().strftime('%Y-%m-%d 00:00')
    last_month_start = time_obj.last_month_start().strftime('%Y-%m-%d 00:00')
    last_month_end = time_obj.last_month_end().strftime('%Y-%m-%d 00:00')
    # this_year_start = time_obj.this_year_start().strftime('%Y-%m-%d 00:00')

    current_year = datetime.now().year
    value_list = ['ownerid','starttime','content','accountid','accountname','opportunityname','longitude','latitude','location','distance','problem_type']
    visit_df = get_value_list(SalesPortraitActivity,
                              {'starttime__startswith':current_year,'entitytype':'4173079'},
                              value_list)
    visit_df = pd.DataFrame(data=visit_df,columns=value_list)
    # print(visit_df)
    # 开始分类计算签到次数
    yestoday_visit = visit_df.loc[visit_df.starttime.str.contains(yestoday)] # 后续需要改成当天的
    # last_hours_visit = visit_df.loc[visit_df.starttime.str.contains(last_hours_start)]
    today_visit = visit_df.loc[visit_df.starttime.str.contains(today)]
    today_count = len(today_visit)
    yestoday_count = len(yestoday_visit)
    this_week_count = len(visit_df.loc[visit_df.starttime>=this_week_start])
    this_month_count = len(visit_df.loc[visit_df.starttime>=this_month_start])
    last_month_count = len(visit_df.loc[(visit_df.starttime>=last_month_start)&(visit_df.starttime<this_month_start)])
    this_year_count = len(visit_df)

    result_df = pd.DataFrame({'date':['今日','昨日','本周','本月','上月','本年度'],
                              'count':[today_count,yestoday_count,this_week_count,this_month_count,last_month_count,this_year_count],
                              'activity_start':[today,yestoday,this_week_start.split(' ')[0],this_month_start.split(' ')[0],last_month_start.split(' ')[0],str(current_year)+'-01-01'],
                              'activity_finish':[today,yestoday,today,today,last_month_end.split(' ')[0],today]
                              })
    # print(today_count)

    # 获取政务各大区的大区名 大区id 所有人员  和ownerid ,使用名字配对，李博的情况会找不到ownerid
    sub_groups = get_value_list(DingGroupMemberMap,{'group_name':'销售中心','sub_group_names__contains':'一大区'},['sub_group_ids'])
    sub_group_ids = json.loads(sub_groups[0][0]) # [58084244, 58087244, 58068265, 155650269, 58111257, 98192107, 98019131, 98006134, 98266229, 98187117, 58040397, 61147353, 111457167]
    group_count_df = get_value_list(DingGroupMemberMap,{'group_id__in':sub_group_ids},['group_name','group_id','member_names_all'])
    group_count_df = pd.DataFrame(data=group_count_df,columns=['group_name','group_id','member_names_all'])
    group_count_df = group_count_df.loc[group_count_df.group_name.str.contains('区|行业拓展')]

    # 特殊处理一些名字对不上的  后台：李博 crm：李博1
    member_names_all = json.loads(list(group_count_df.loc[group_count_df.group_name.str.contains('上海'),'member_names_all'])[0])
    # print('修改前',member_names_all)
    for i in range(len(member_names_all)):
        if member_names_all[i] == '李博':
            member_names_all[i] = '李博1'
    member_names_all = json.dumps(member_names_all)
    group_count_df.loc[group_count_df.group_name.str.contains('上海'), 'member_names_all'] = member_names_all
    member_names_all = json.loads(list(group_count_df.loc[group_count_df.group_name.str.contains('上海'),'member_names_all'])[0])
    # print('修改后',member_names_all)

    group_count_df['ownerids'] = group_count_df['member_names_all'].apply(lambda x:list(CrmUser.objects.filter(name__in=json.loads(x)).values_list("id", flat=True)))
    yestoday_visit.loc[:, 'ownerid'] = yestoday_visit.loc[:, 'ownerid'].astype(np.int64)
    group_count_df['count'] = group_count_df['ownerids'].apply(
        lambda x: len(yestoday_visit.loc[yestoday_visit.ownerid.isin(x)]))  # 计算大区的拜访次数
    group_count_df = group_count_df[['group_id','group_name', 'count']].sort_values(by='count', ascending=False) # 排序

    group_count_df['num'] = [i for i in range(1,len(group_count_df)+1)]  #处理排名序号
    group_count_df['activity_start'] = yestoday
    group_count_df['activity_finish'] = yestoday

    # 转换yestoday_visit 的ownerid列的格式由object 为np.int64
    # 当天的签到具体数据
    if len(today_visit)>0:
        owner_df = get_value_list(CrmUser,{'id__in':list(set(today_visit.loc[:,"ownerid"]))},['id','name'])
        owner_df = pd.DataFrame(data=owner_df,columns=['ownerid','salename'])
        owner_df['ownerid'] = owner_df['ownerid'].astype(str)
        today_visit = today_visit.sort_values(by='starttime',ascending=False)
        today_visit = pd.merge(today_visit,owner_df,how='left',on='ownerid')
        today_visit = today_visit.drop(['ownerid'],axis=1)
        # print(list(today_visit['problem_type']))
        today_visit['problem_content'] = today_visit['problem_type'].apply(lambda x:problem_content.get(x))
        # print(list(today_visit['problem_content']))
        today_visit = today_visit.fillna('')
        # print(today_visit)
        visit_detail = list(today_visit.to_dict(orient="index").values())
    else:
        visit_detail = []


    # 大区排名数据,应该用今日的数据
    group_count = list(group_count_df.to_dict(orient="index").values())
    # 首页的签到次数
    visit_count = list(result_df.to_dict(orient="index").values())
    # 地图和签到详情数据 visit_detail
    result['visit_detail'] = visit_detail
    result['visit_count'] = visit_count
    result['group_count'] = group_count
    result['end_time'] = end_time

    return result



'''签到选择页面，条件为商务 部门 选择的签到开始日和结束日 '''
@login_required()
@permission_required("salerVisit.select_visit.view",login_url=URL_403)
def selectVisitHtml(request):
    form = VisitSelectForm()
    departments_api = reverse("secretary:departments_api")
    department_data = get_departments_about_sale()
    department_data = json.dumps(department_data)
    department_data_1 = [["", "全部"]] + list(settings.SALE_DEPARTMENT_LEVEL_1.items())
    department_data_1 = json.dumps(department_data_1)
    saler = request.GET.get("saler")
    department_id = request.GET.get("department")
    activity_start = request.GET.get("activity_start")
    activity_finish = request.GET.get("activity_finish")
    if saler:
        form.input_saler(saler)
    if department_id:
        form.select_department(department_id)
    if activity_start:
        form.input_activity_start(activity_start)
    if activity_finish:
        form.input_activity_finish(activity_finish)


    title = "签到筛选页面"
    user = request.user
    if user.username not in ["武翠霞", "丛大侠", "王寅", "张太锋", "黄远", "朱铭鑫", "袁战梅", "牛于斌"]:
        message = json.dumps(dict(request.GET))
        UserLog.objects.create(username=user.username, user_id=user.id,
                               model="签到筛选", action="进入页面", message=message)
    return render(request, 'salerVisit/select_visit.html', locals())


'''处理签到筛选页面的筛选数据接口'''
@login_required()
@permission_required("salerVisit.visit_detail.view",login_url=JSON_403)
def selectVisitApi(request):
    result = selectVisitMethod(request)
    return JsonResponse(result)


def selectVisitMethod(request):
    result = {"status": 0, "message": "", "items": []}
    saler = request.GET.get("saler")
    department_id = request.GET.get("department")
    activity_start = request.GET.get("activity_start")
    activity_finish = request.GET.get("activity_finish")
    problem_type = request.GET.get("problem_type")
    access_type = request.GET.get("type") # 请求数据来源，如果是详情列表detail_table 则增加个人签到次数排名数据返回

    # print(saler, department_id, activity_start, activity_finish)


    select_condition = {}
    ownerlist = []
    owner_df = pd.DataFrame(columns=['ownerid', 'salename','employeecode']) # 商务的id和名字对应表
    # 按照商务条件筛选
    if saler:
        owner_df = pd.DataFrame(data=get_value_list(CrmUser, {"name__contains": saler}, ["id","name","employeecode"]),
                                columns=['ownerid', 'salename','istarshine_id'])

        ownerlist = list(owner_df["ownerid"])
        if len(ownerlist) == 0:
            result["message"] = "未查到该商务信息"
            return result

    # 按部门条件筛选
    if department_id:
        group_members = get_group_members(department_id)
        if not group_members:
            result["message"] = "部门成员为空，请检查部门是否存在"
            return result
        else:
            group_owner_df = pd.DataFrame(data=get_value_list(CrmUser, {"name__in": group_members}, ["id","name","employeecode"]),
                                          columns=['ownerid', 'salename','istarshine_id'])
            group_members_owner = list(group_owner_df["ownerid"])
            if group_members_owner:
                if ownerlist:
                    ownerlist = list(set(ownerlist) & set(group_members_owner))
                    if not ownerlist:
                        result["message"] = "搜索的商务不在该部门下，请修改部门查询"
                        return result
                    else:
                        owner_df = owner_df.loc[owner_df.ownerid.isin(ownerlist)]
                else:
                    ownerlist += group_members_owner
                    owner_df = owner_df.append(group_owner_df, ignore_index=False, sort=None)
            else:
                result["message"] = "部门成员下无信息，请修改部门查询"
                return result

    if ownerlist:
        select_condition["ownerid__in"] = ownerlist

    # 签到开始日期
    if activity_start:
        select_condition["starttime__gte"] = activity_start + ' 00:00'

    # 签到结束日期
    if activity_finish:
        select_condition["starttime__lte"] = activity_finish + ' 23:59'

    if not activity_start and not activity_finish:
        current_year = datetime.now().year
        select_condition["starttime__startswith"] = current_year

    if problem_type:
        if problem_type == '0':
            select_condition["problem_type"] = None
        else:
            select_condition["problem_type"] = problem_type

    select_condition["entitytype"] = "4173079"  # 签到拜访

    # print(select_condition)
    value_list = ['activityid','ownerid', 'starttime','content', 'accountname', 'opportunityname', 'longitude', 'latitude', 'location',
                  'distance','problem_type']
    visit_df = get_value_list(SalesPortraitActivity,
                              select_condition,
                              value_list)
    visit_df = pd.DataFrame(data=visit_df, columns=value_list)
    if len(owner_df)==0:
        owner_df = pd.DataFrame(data=get_value_list(CrmUser, {"id__in": list(visit_df["ownerid"])}, ["id", "name","employeecode"]),
                                columns=['ownerid', 'salename','istarshine_id'])

    # 获取商务 - 部门对应关系
    user_group_df = get_groups(list(owner_df['istarshine_id']))

    owner_df = pd.merge(owner_df,user_group_df,how='left',on='istarshine_id')

    owner_df['ownerid'] = owner_df['ownerid'].astype(str)


    visit_df = visit_df.sort_values(by='starttime', ascending=True)
    visit_df = pd.merge(visit_df, owner_df, how='left', on='ownerid')
    visit_df = visit_df.drop(['ownerid'], axis=1)
    # 获取签到问题描述
    visit_df['problem_content'] = visit_df['problem_type'].apply(lambda x: problem_content.get(x))
    problem_df = pd.DataFrame(get_value_list(NoticeList,{'notice_type':12,'object_type':5},['object_id','content']),
                              columns=['activityid','problem_content'])
    problem_df.drop_duplicates(subset=['activityid'], keep="first", inplace=True)

    for index,row in problem_df.iterrows():
        # print(row['problem_content'])
        if row['problem_content']:
            visit_df.loc[visit_df.activityid==str(row.activityid),'problem_content'] = '初步判断：' + row['problem_content'].split('初步判断：')[-1]
        else:
            print(row['problem_content'])
            visit_df.loc[visit_df.activityid == str(row.activityid), 'problem_content'] = ''

    # print(list(visit_df['problem_content']))


    visit_df = visit_df.fillna("")
    visit_df['index'] = visit_df.index + 1
    result['items'] = list(visit_df.to_dict(orient="index").values())

    if access_type == 'detail_table':
        # visit_df.to_csv('签到拜访情况统计.csv', index=False, encoding="utf_8_sig")
        # 整理个人签到排序数据
        staff_visit_count = pd.DataFrame(columns=['salename','count'])
        for staff in set(visit_df['salename']):
            count = len(visit_df.loc[visit_df.salename==staff])
            staff_visit_count.loc[len(staff_visit_count)] = [staff,count]
        staff_visit_count = staff_visit_count.sort_values(by='count',ascending=False)
        result['staff_visit_count'] = list(staff_visit_count.to_dict(orient='index').values())

    result['message'] = 'success'
    result['status'] = 1

    return result

@login_required()
@permission_required("salerVisit_visit_table.export",login_url=URL_403)
def export_excel(request):
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = '签到拜访信息'  # 给工作表1设置标题
    row_one = ['签到记录时间','拜访内容','最终客户','商机','签到经度','签到维度','签到地址','签到与客户距离','商务','部门','上级部门','签到问题']

    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    all_obj = selectVisitMethod(request=request).get('items',[])
    for obj in all_obj:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        # print(obj)
        obj_info = [obj['starttime'],obj['content'],obj['accountname'],obj['opportunityname'],
                    obj['longitude'],obj['latitude'],obj['location'],obj['distance'],obj['salename'],
                    obj['departments'],obj['departments_top'],obj['problem_content']]
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]

    output_io = BytesIO()
    wb.save(output_io)  # 将Excel文件内容保存到IO中
    output_io.seek(0)  # 重新定位到开始

    # 设置HttpResponse的类型
    response = HttpResponse(output_io.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    file_name = 'VisitDetails_%s.xls' % ctime  # 给文件名中添加日期时间
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name

    user = request.user
    if user.username not in ["武翠霞", "丛大侠", "王寅", "张太锋", "黄远", "朱铭鑫", "袁战梅", "牛于斌"]:
        message = json.dumps(dict(request.GET))
        UserLog.objects.create(username=user.username, user_id=user.id,
                               model="签到数据详情", action="导出", message=message)
    return response


'''签到列表详情页面，条件为商务 部门 选择的签到开始日和结束日 '''
@login_required()
@permission_required("salerVisit_visit_table.view",login_url=URL_403)
def VisitTableHtml(request):
    form = VisitSelectForm()
    saler = request.GET.get("saler")
    department_id = request.GET.get("department")
    activity_start = request.GET.get("activity_start")
    activity_finish = request.GET.get("activity_finish")
    if saler:
        form.input_saler(saler)
    if department_id:
        form.select_department(department_id)

        top_department_df = get_department([department_id,])

    if activity_start:
        form.input_activity_start(activity_start)
    if activity_finish:
        form.input_activity_finish(activity_finish)

    departments_api = reverse("secretary:departments_api")
    department_data = get_departments_about_sale()
    department_data = json.dumps(department_data)
    department_data_1 = [["", "全部"]] + list(settings.SALE_DEPARTMENT_LEVEL_1.items())
    department_data_1 = json.dumps(department_data_1)

    init_data = {"row_list": [10, 20, 50], "row_num": 20}
    init_data["col_names"] = ["序号","商务","部门","上级部门","签到问题", "签到时间","签到地", "签到距离(公里)", "最终客户", "商机名称", "沟通内容"]
    init_data["col_model"] = [
        {
            "name": 'index',
            "width": 5,
            "sorttype": "number"
        },
        {
            "name": 'salename',
            "width": 8,
        },
        {
            "name": 'departments',
            "width": 10,
        },
        {
            "name": 'departments_top',
            "width": 10,
        },
        {
            "name": 'problem_content',
            "width": 15,
        },
        {
            "name": 'starttime',
            "width": 15,
        },
        {
            "name": 'location',
            "width": 25,
        },
        {
            "name": 'distance',
            "width": 10,
            "sorttype": "float"
        },
        {
            "name": 'accountname',
            "width": 20,
        },
        {
            "name": 'opportunityname',
            "width": 20,
        },
        {
            "name": 'content',
            "width": 30,
        }

        # {
        #     "name": 'operations',
        #     "width": 10,
        #     "sortable": False,
        # },
    ]
    init_data["items"] = []
    init_data = json.dumps(init_data)

    # problem_type = [('','全部')] + [(k,v) for k,v in problem_content1.items()]

    title = "签到数据详情"
    user = request.user
    if user.username not in ["武翠霞", "丛大侠", "王寅", "张太锋", "黄远", "朱铭鑫", "袁战梅"]:
        message = json.dumps(dict(request.GET))
        UserLog.objects.create(username=user.username, user_id=user.id,
                               model="签到数据详情", action="进入页面", message=message)
    return render(request, 'salerVisit/visit_detail_list.html', locals())
